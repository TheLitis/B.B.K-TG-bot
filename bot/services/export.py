"""Export utilities for selections and requests."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from jinja2 import Environment, StrictUndefined
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

env = Environment(undefined=StrictUndefined, trim_blocks=True, lstrip_blocks=True)


@dataclass(slots=True)
class SelectionLine:
    """Single item in the selection export."""

    sku: str
    name: str
    category: str
    brand: str
    area_m2: float
    waste_pct: int
    total_m2: float
    pack_step: float | None = None
    notes: str | None = None


def selection_to_workbook(
    items: Iterable[SelectionLine],
    customer: dict[str, Any],
    company: dict[str, Any] | None = None,
) -> BytesIO:
    """Generate Excel workbook with selection details."""

    items_list = list(items)
    workbook = Workbook()
    ws_items = workbook.active
    ws_items.title = "Подборка"

    headers = [
        "SKU",
        "Категория",
        "Коллекция",
        "Бренд",
        "Площадь, м²",
        "Запас, %",
        "Итого, м²",
        "Кратность упаковки",
        "Комментарии",
    ]
    ws_items.append(headers)
    header_font = Font(bold=True)
    for cell in ws_items[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    totals = 0.0
    for line in items_list:
        ws_items.append(
            [
                line.sku,
                line.category,
                line.name,
                line.brand,
                round(line.area_m2, 2),
                line.waste_pct,
                round(line.total_m2, 2),
                line.pack_step if line.pack_step else "",
                line.notes or "",
            ]
        )
        totals += line.total_m2

    # Totals sheet ----------------------------------------------------------------
    ws_summary = workbook.create_sheet("Итоги")
    ws_summary.append(["Всего позиций", len(items_list)])
    ws_summary.append(["Общий метраж, м²", round(totals, 2)])

    # Contacts sheet --------------------------------------------------------------
    ws_contacts = workbook.create_sheet("Контакты клиента")
    for key, value in customer.items():
        ws_contacts.append([key, value])

    if company:
        ws_contacts.append([])
        ws_contacts.append(["Контакты компании"])
        for key, value in company.items():
            ws_contacts.append([key, value])

    auto_fit_columns(ws_items)
    auto_fit_columns(ws_summary)
    auto_fit_columns(ws_contacts)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def auto_fit_columns(sheet) -> None:
    """Adjust sheet columns based on their content length."""

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column].width = max_length + 2


def render_html_request(selection: list[SelectionLine], customer: dict[str, Any]) -> str:
    """Render HTML summary for selections (can be converted to PDF later)."""

    template = env.from_string(
        """
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; color: #222; }
                table { border-collapse: collapse; width: 100%; margin-top: 16px; }
                th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
                th { background-color: #f4f4f4; }
            </style>
        </head>
        <body>
            <h2>Заявка на просчёт / образцы</h2>
            <p>Дата: {{ now }}</p>
            <h3>Клиент</h3>
            <ul>
            {%- for key, value in customer.items() %}
                <li><strong>{{ key }}:</strong> {{ value }}</li>
            {%- endfor %}
            </ul>
            <h3>Подборка</h3>
            <table>
                <thead>
                    <tr>
                        <th>SKU</th>
                        <th>Категория</th>
                        <th>Коллекция</th>
                        <th>Бренд</th>
                        <th>Площадь, м²</th>
                        <th>Запас, %</th>
                        <th>Итого, м²</th>
                    </tr>
                </thead>
                <tbody>
                {% for line in selection %}
                    <tr>
                        <td>{{ line.sku }}</td>
                        <td>{{ line.category }}</td>
                        <td>{{ line.name }}</td>
                        <td>{{ line.brand }}</td>
                        <td>{{ "%.2f"|format(line.area_m2) }}</td>
                        <td>{{ line.waste_pct }}</td>
                        <td>{{ "%.2f"|format(line.total_m2) }}</td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </body>
        </html>
        """
    )
    from datetime import datetime

    return template.render(
        selection=selection,
        customer=customer,
        now=datetime.now().strftime("%d.%m.%Y %H:%M"),
    )


__all__ = ["SelectionLine", "selection_to_workbook", "render_html_request"]
